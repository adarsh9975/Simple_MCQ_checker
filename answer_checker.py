'''####################################################################
## Author: Adarsh Sinha                                              ##
## Date: 25 Sep 2019                                                 ##
## Project: Simple_MCQ_Checker                                       ##
## Python version: 3.x                                               ##
## License: MIT                                                      ##
## Description: Checks all the assessment questionnaires using the   ##
##              provided answer key and generates a report in excel  ##
##              containing Name,ID and score of the candidate.       ##
#####################################################################'''
import openpyxl as xl
import glob

PATH_TO_ANSWER_SHEET ='./Assess_answer.xlsx'

def answer_checker(PATH_TO_ASSESMENT):
    '''********** Open Excel file to check ***********'''
    to_check_work_book= xl.load_workbook(PATH_TO_ASSESMENT);
    '''************** Open Answer file ***************'''
    answer_book= xl.load_workbook(PATH_TO_ANSWER_SHEET);
    '''************ Open Sheet to check **************'''
    to_check_sheet= to_check_work_book['Sheet1'];
    '''************** Open Answer Sheet **************'''
    answer_sheet= answer_book['Sheet1'];

    '''***************** Access Name *****************'''
    candidate_name= to_check_sheet['A2'].value;
    print('Candidate Name is: ',candidate_name);

    '''************** Access Employee ID *************'''
    candidate_id= to_check_sheet['C2'].value;
    print('Candidate ID is: ',candidate_id);

    '''************ Access candidate answers **********'''
    candidate_answers_col=to_check_sheet['C'];
    candidate_answer_list=[];
    iterator=0;
    while (iterator<len(candidate_answers_col))  :
        if (((candidate_answers_col[iterator].value)=='a.') or \
        ((candidate_answers_col[iterator].value)=='b.') or \
        ((candidate_answers_col[iterator].value)=='c.') or \
        ((candidate_answers_col[iterator].value)=='d.')):
            '''capture question number (is one row above currrent row )'''
            question_num=to_check_sheet.cell(row=iterator,column=1).value;
            '''capture answer'''
            answer=candidate_answers_col[iterator].value;
            '''store in a list'''
            candidate_answer_list.append((question_num,answer));
        iterator+=1;

    #print('Candidate Answers: ',candidate_answer_list)

    '''********* Access answers in answer sheet ******'''
    true_answer_list=answer_sheet['A'];

    '''***************** Check answers ***************'''
    iterator =0;
    score=0;
    #print('List_len: ',len(candidate_answer_list))
    while (iterator<len(candidate_answer_list)):
        question_number=int(candidate_answer_list[iterator][0][1:]);
        #print('Assessing Question number : ',question_number)
        if (candidate_answer_list[iterator][1] == true_answer_list[question_number-1].value):
            print('Question',question_number,' Answered correctly');
            score+=1;
        else:
            print('Question',question_number,' Answered wrong');
        iterator+=1;

    '''****************** Print score ****************'''
    print(candidate_name,'(',candidate_id,')','has a final score of: ',score);
    print()#for asthetics
    print()#for asthetics
    return(candidate_name,candidate_id,score)


def main():
    '''****Extended functionality to multiple input assessment files ****'''
    iterator=0;
    write_iterator=0;
    '''****Create a workbook to save the result to****'''
    result_workbook= xl.Workbook();
    result_worksheet= result_workbook.create_sheet('Result',0);
    '''*******************Write Titles in Row 1**************************'''
    result_worksheet[('A'+str(1))].value = 'Candidate Name';
    result_worksheet[('B'+str(1))].value = 'Candidate ID';
    result_worksheet[('C'+str(1))].value = 'Score';

    '''**** Get all the excel files in the folder ****'''
    List_of_all_sheet_to_assess= glob.glob("./*.xlsx");

    '''**** Check each one by one (except the answersheet itself) ****'''
    while(iterator<len(List_of_all_sheet_to_assess)):
        if (List_of_all_sheet_to_assess[iterator]!=PATH_TO_ANSWER_SHEET):
            result=answer_checker(List_of_all_sheet_to_assess[iterator])
            #print(result)
            '''******* Save result to an excel workbook ******'''
            '''*Start writing from row 2 since row 1 has titles (hence iterator+2)*'''
            #write name
            result_worksheet[('A'+str(write_iterator+2))].value = result[0];
            #write candidate id
            result_worksheet[('B'+str(write_iterator+2))].value = result[1];
            #write score
            result_worksheet[('C'+str(write_iterator+2))].value = result[2];
            write_iterator+=1;
        iterator+=1;
    '''*****************Save the consolidated result******************'''
    result_workbook.save('Consolidated_result.xlsx');

if __name__== "__main__":
  main()
